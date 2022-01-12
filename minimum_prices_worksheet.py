import logging
from operator import attrgetter
from typing import NamedTuple
import openpyxl


class SalesOrdersRow(NamedTuple):
    OrderDate: str
    Region: str
    Rep: str
    Item: str
    Units: str
    UnitCost: str
    Total: str


def read_data_worksheet(inwsheet) -> list[SalesOrdersRow]:
    listRows = []
    for i in range(2, inwsheet.max_row + 1):
        listRows.append(SalesOrdersRow(inwsheet.cell(row=i, column=1).value, inwsheet.cell(row=i, column=2).value,
                                       inwsheet.cell(row=i, column=3).value, inwsheet.cell(row=i, column=4).value,
                                       inwsheet.cell(row=i, column=5).value, inwsheet.cell(row=i, column=6).value,
                                       inwsheet.cell(row=i, column=7).value))
    return listRows


def filter_data(data: list[SalesOrdersRow]) -> list[SalesOrdersRow]:
    sortedListRows = sorted(data, key=attrgetter('Item', 'UnitCost'))

    minPriceList = []
    dictsheet = {}
    for orderRow in sortedListRows:
        if orderRow.Item not in dictsheet:
            dictsheet[orderRow.Item] = orderRow.UnitCost
            minPriceList.append(orderRow)
            logging.info(f'Minimum price for {orderRow.Item} is {orderRow.UnitCost}')
        else:
            if dictsheet[orderRow.Item] == orderRow.UnitCost:
                minPriceList.append(orderRow)
                logging.info(
                    f'Another {orderRow.Item} from the sheet has minimum price {orderRow.UnitCost}')
    return minPriceList


def write_data(wsheet, wsheet2, data: list[SalesOrdersRow]):
    for l in range(1, wsheet.max_column + 1):
        wsheet2.cell(row=1, column=l).value = wsheet.cell(row=1, column=l).value
    logging.info('First row was copied in the new sheet')
    for i in range(2, len(data) + 2):
        for j in range(1, 8):
            wsheet2.cell(row=i, column=j).value = data[i - 2][j - 1]
        logging.info(f'Row {i} was copied in the new sheet.')

"""
Find the items with the lowest price by Item and UnitCost in 'SalesOrders' worksheet 
Copies this info in 'minimum prices' sheet.
"""
logging.basicConfig(level=logging.INFO)

wb = openpyxl.load_workbook('SampleData.xlsx')
wsheet = wb.get_sheet_by_name('SalesOrders')

listRows = read_data_worksheet(wsheet)
newsheetList = filter_data(listRows)
wsheetMinPrice = wb.create_sheet()
wsheetMinPrice.title = 'minimum prices'
write_data(wsheet, wsheetMinPrice, newsheetList)
wb.save('SampleData.xlsx')